import openpyxl

class HomePageData:

    @staticmethod
    def getTestData():
        test_data = []

        # loading workbook
        book = openpyxl.load_workbook("C:\\Users\\ckula\\PycharmProjects\\PythonTestingFramework\\test_data\\test_data.xlsx")
        # fetching active sheet
        sheet = book.active

        # fetching all the data

        for row in range(2, sheet.max_row + 1):
            row_data = []
            for col in range(1, sheet.max_column + 1):
                row_data.append(sheet.cell(row=row, column=col).value)

            test_data.append(tuple(row_data))

        return test_data